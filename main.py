import pandas as pd
import requests
from lxml import html
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# 定义 Anime 类，并增加更多的属性用于存储不同平台的数据
class Anime:
    def __init__(self, original_name, score_bgm='', score_al='', score_mal='', score_fm='',
                 bangumi_url='', anilist_url='', myanilist_url='', filmarks_url='',
                 bangumi_name='', anilist_name='', myanilist_name='', flimarks_name=''):
        self.original_name = original_name  # 原始名称
        self.score_bgm = score_bgm  # Bangumi评分
        self.score_al = score_al  # AniList评分
        self.score_mal = score_mal  # MyAnimeList评分
        self.score_fm = score_fm  # Filmarks评分
        self.bangumi_url = bangumi_url  # Bangumi条目链接
        self.anilist_url = anilist_url  # AniList条目链接
        self.myanilist_url = myanilist_url  # MyAnimeList条目链接
        self.filmarks_url = filmarks_url  # Filmarks条目链接
        self.bangumi_name = bangumi_name  # Bangumi名称
        self.anilist_name = anilist_name  # AniList名称
        self.myanilist_name = myanilist_name  # MyAnimeList名称
        self.flimarks_name = flimarks_name  # Filmarks名称

    def __str__(self):
        return (f"Anime({self.original_name}, BGM: {self.score_bgm}, AL: {self.score_al}, "
                f"MAL: {self.score_mal}, FM: {self.score_fm}, "
                f"URLs: {self.bangumi_url}, {self.anilist_url}, {self.myanilist_url}, {self.filmarks_url}, "
                f"Names: {self.bangumi_name}, {self.anilist_name}, {self.myanilist_name}, {self.flimarks_name})")


# 读取Excel文件，假设文件名为test.xlsx
file_path = 'test2.xlsx'
df = pd.read_excel(file_path)

# 创建一个空的anime_list列表，用于存储Anime类的实例
anime_list = []

# 遍历DataFrame中的每一行数据
for index, row in df.iterrows():
    if pd.isna(row['原名']):
        print(f"Skipping row {index} because the original name is NaN.")
        continue  # 跳过这一行
    new_anime = Anime(original_name=row['原名'])  # 获取每行的“原名”列作为原始名称
    anime_list.append(new_anime)  # 将新的Anime实例加入列表

# 定义 AniList GraphQL API 的URL
anilist_url = 'https://graphql.anilist.co'

# 遍历anime_list列表
for anime in anime_list:
    print(anime)
    # URL编码处理动漫名称，避免特殊字符带来的问题
    keyword_encoded = requests.utils.quote(anime.original_name.encode('utf-8'))


    # 1. 调用 Bangumi API 搜索条目
    search_url = f"https://api.bgm.tv/search/subject/{keyword_encoded}"
    search_params = {
        'responseGroup': 'small',
        'type': 2  # 假设我们只关注动画类型的条目
    }
    search_response = requests.get(search_url, params=search_params)
    search_data = search_response.json()

    # 如果搜索结果包含列表且有条目
    if 'list' in search_data and search_data['list']:
        first_subject_id = search_data['list'][0]['id']  # 获取第一个条目的ID
        anime.bangumi_url = f"https://bgm.tv/subject/{first_subject_id}"  # 构造Bangumi条目URL
        anime.bangumi_name = search_data['list'][0].get('name', 'No name found')  # 获取Bangumi名称

        # 请求获取详细条目数据
        subject_url = f"https://api.bgm.tv/subject/{first_subject_id}"
        subject_response = requests.get(subject_url)
        subject_data = subject_response.json()

        # 如果存在评分信息，则保存评分，否则标记为未找到评分
        if 'rating' in subject_data and 'score' in subject_data['rating']:
            anime.score_bgm = subject_data['rating']['score']
        else:
            anime.score_bgm = 'No score available'
        print("bgm的评分"+str(anime.score_bgm))
        print("bgm的条目名字"+str(anime.bangumi_name))
        print("bgm的链接"+str(anime.bangumi_url))
    else:
        anime.score_bgm = 'No results found'  # 没有搜索到结果

    # 2. 调用 MyAnimeList 页面搜索
    mal_search_url = f"https://myanimelist.net/anime.php?q={keyword_encoded}&cat=anime"
    mal_search_response = requests.get(mal_search_url)

    if mal_search_response.status_code == 200:
        mal_tree = html.fromstring(mal_search_response.content)

        try:
            # 获取搜索结果中第一个条目的链接
            anime_href = mal_tree.xpath('/html/body/div[1]/div[2]/div[4]/div[2]/div[6]/table/tr[2]/td[2]/div[1]/a[1]/@href')[0]
            anime.myanilist_url = anime_href  # 存储MyAnimeList条目链接
            mal_anime_response = requests.get(anime_href)

            if mal_anime_response.status_code == 200:
                mal_anime_tree = html.fromstring(mal_anime_response.content)

                # 获取评分信息
                anime_mal_score = mal_anime_tree.xpath(
                    '/html/body/div[1]/div[2]/div[4]/div[2]/table/tr[1]/td[2]/div[1]/table/tr[1]/td/div[1]/div[1]/div[1]/div[1]/div[1]/div/text()')
                anime_mal_score = anime_mal_score[0].strip() if anime_mal_score else 'No score found'
                anime.score_mal = anime_mal_score  # 保存评分

                # 获取MyAnimeList的名称
                myanimelist_name = mal_anime_tree.xpath(
                    '/html/body/div[1]/div[2]/div[4]/div[1]/div/div[1]/div/h1/strong/text()')
                anime.myanilist_name = myanimelist_name[0].strip() if myanimelist_name else 'No name found'
            else:
                anime.score_mal = 'No score found'
            print("mal的评分"+str(anime.score_mal))
            print("mal的名称"+str(anime.myanilist_name))
            print("mal的链接"+str(anime.myanilist_url))
        except IndexError:
            anime.score_mal = 'No href found'  # 没有找到条目链接
    else:
        anime.score_mal = 'No results found'  # 请求失败

    # 3. 调用 AniList API 搜索
    anilist_search_query = '''
    query ($search: String) {
      Page (page: 1, perPage: 1) {
        media (search: $search, type: ANIME) {
          id
          title {
            romaji
          }
        }
      }
    }
    '''
    anilist_search_variables = {
        "search": anime.original_name  # 使用原始名称进行搜索
    }

    anilist_search_response = requests.post(anilist_url,
                                            json={'query': anilist_search_query, 'variables': anilist_search_variables})
    anilist_search_data = anilist_search_response.json()

    # 处理AniList API返回的数据
    if 'data' in anilist_search_data and 'Page' in anilist_search_data['data']:
        media_list = anilist_search_data['data']['Page']['media']
        if media_list:
            first_anime_id = media_list[0]['id']  # 获取第一个条目的ID
            anime.anilist_url = f"https://anilist.co/anime/{first_anime_id}"  # 构造AniList条目URL
            anime.anilist_name = media_list[0]['title']['romaji']  # 获取AniList名称

            # 请求获取详细评分信息
            anilist_detail_query = '''
            query ($id: Int) {
              Media (id: $id) {
                averageScore
              }
            }
            '''
            anilist_detail_variables = {
                "id": first_anime_id
            }

            anilist_detail_response = requests.post(anilist_url, json={'query': anilist_detail_query,
                                                                       'variables': anilist_detail_variables})
            anilist_detail_data = anilist_detail_response.json()

            # 如果找到评分信息，则存储
            if 'data' in anilist_detail_data and 'Media' in anilist_detail_data['data']:
                anime_detail = anilist_detail_data['data']['Media']
                anime.score_al = anime_detail.get('averageScore', 'No score found')
        else:
            anime.score_al = 'No AniList results'  # 没有找到条目
        print("al的评分"+str(anime.score_al))
        print("al的名称"+str(anime.anilist_name))
        print("al的链接"+str(anime.anilist_url))
    else:
        anime.score_al = 'Error with AniList API'  # API请求出错

    # 4. 调用 Filmarks API 搜索
    filmarks_url = f"https://filmarks.com/search/animes?q={keyword_encoded}"
    filmarks_response = requests.get(filmarks_url)

    if filmarks_response.status_code == 200:
        filmarks_tree = html.fromstring(filmarks_response.content)
        try:
            # 获取评分信息
            anime_fm_score = filmarks_tree.xpath(
                '/html/body/div[3]/div[3]/div[2]/div[1]/div[2]/div/div[1]/div[2]/div[3]/div/div[2]/text()')
            anime.score_fm = anime_fm_score[0].strip() if anime_fm_score else 'No score found'
            anime.filmarks_url = filmarks_url  # 存储Filmarks URL

            # 获取Filmarks名称
            filmarks_name = filmarks_tree.xpath(
                '/html/body/div[3]/div[3]/div[2]/div[1]/div[2]/div/div[1]/div[2]/div[1]/h3/text()')
            anime.flimarks_name = filmarks_name[0].strip() if filmarks_name else 'No name found'

            print("fm的评分"+str(anime.score_fm))
            print("fm的名字"+str(anime.flimarks_name))
            print("fm的链接"+str(anime.filmarks_url))
        except IndexError:
            anime.score_fm = 'No Filmarks score found'  # 没有找到评分
    else:
        anime.score_fm = 'No Filmarks results'  # Filmarks请求失败

# 打印anime_list中的所有条目，验证各个属性
# for anime in anime_list:
#     print(f"Original Name: {anime.original_name}, Score BGM: {anime.score_bgm}, "
#           f"Score MAL: {anime.score_mal}, Score AL: {anime.score_al}, Score FM: {anime.score_fm}, "
#           f"Bangumi URL: {anime.bangumi_url}, AniList URL: {anime.anilist_url}, "
#           f"MyAnimeList URL: {anime.myanilist_url}, Filmarks URL: {anime.filmarks_url}, "
#           f"Bangumi Name: {anime.bangumi_name}, AniList Name: {anime.anilist_name}, "
#           f"MyAnimeList Name: {anime.myanilist_name}, Filmarks Name: {anime.flimarks_name}")

# 将数据写回Excel文件
wb = load_workbook(file_path)
ws = wb.active

# 定义一个样式，用于设置单元格格式为数值
number_style = NamedStyle(name="number_style", number_format='0.00')

# 遍历anime_list，将数据写入对应的Excel单元格
for anime in anime_list:
    for row in ws.iter_rows(min_row=2):  # 假设第一行为表头，从第2行开始遍历
        if row[0].value == anime.original_name:  # 匹配原始名称
            # 写入评分，并设置数值格式
            row[2].value = float(anime.score_bgm) if anime.score_bgm not in ['No score available', 'No results found', None] else None
            # row[2].number_format = 'General'

            row[3].value = float(anime.score_al) / 10 if anime.score_al not in ['No score found', 'No AniList results', None] else None
            # row[3].number_format = 'General'

            row[4].value = float(anime.score_mal) if anime.score_mal not in ['No score found', None] else None
            # row[4].number_format = 'General'

            row[5].value = float(anime.score_fm)  if anime.score_fm not in ['No score found', 'No Filmarks results', None] else None
            # a = row[5].value
            # print(a)
            # row[5].number_format = 'General'

            row[6].value = float(anime.score_fm)*2  if anime.score_fm not in ['No score found', 'No Filmarks results', None] else None
            # a = row[5].value
            # print(a)
            # row[6].number_format = 'General'

            # 写入名称并设置超链接
            row[9].value = anime.bangumi_name if anime.bangumi_name not in ['No name found', None] else None
            if anime.bangumi_url:
                ws.cell(row=row[0].row, column=10).hyperlink = anime.bangumi_url

            row[10].value = anime.anilist_name if anime.anilist_name not in ['No name found', None] else None
            if anime.anilist_url:
                ws.cell(row=row[0].row, column=11).hyperlink = anime.anilist_url

            row[11].value = anime.myanilist_name if anime.myanilist_name not in ['No name found', None] else None
            if anime.myanilist_url:
                ws.cell(row=row[0].row, column=12).hyperlink = anime.myanilist_url

            row[12].value = anime.flimarks_name if anime.flimarks_name not in ['No name found', None] else None
            if anime.filmarks_url:
                ws.cell(row=row[0].row, column=13).hyperlink = anime.filmarks_url


# 保存更新后的Excel文件
wb.save(file_path)

print("Excel表格已成功更新。")

# 保存更新后的Excel文件
wb.save(file_path)

print("Excel表格已成功更新。")

# 循环等待用户输入 'exit' 后退出
while True:
    user_input = input("输入 'exit' 退出程序: ")
    if user_input.lower() == 'exit':  # 忽略大小写，允许 'exit' 退出
        break

print("程序已退出...")
